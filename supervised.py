
import os
# Force Keras on CPU
# os.environ['CUDA_VISIBLE_DEVICES'] = '-1'

import cv2     # for capturing videos
import math   # for mathematical operations
import matplotlib.pyplot as plt    # for plotting the images
import pandas as pd
from keras.preprocessing import image   # for preprocessing the images
import numpy as np    # for mathematical operations
from keras.utils import np_utils
from skimage.transform import resize   # for resizing images
from openpyxl import load_workbook

from keras.layers import TimeDistributed, GRU, Dense, Dropout, LayerNormalization, ConvLSTM2D

from keras.layers import Conv2D, BatchNormalization, \
    MaxPool2D, GlobalMaxPool2D

import keras

import tensorflow as tf
print(tf.test.is_built_with_cuda())
print(tf.config.list_physical_devices())
# gpu_options = tf.GPUOptions(allow_growth=True)
# session = tf.InteractiveSession(config=tf.ConfigProto(gpu_options=gpu_options))

NBFRAME = 5
SIZE = (112, 112)
CHANNELS = 3 # 1 if greyscale

def defineTrainTestSets(NBFRAME = NBFRAME, SIZE = SIZE, CHANNELS=CHANNELS):
    trainList = []
    testList = []

    anomaly_path = 'anomaly/'
    normal_path = 'normal/'
    count=0
    for entry in os.listdir(anomaly_path):
        path = os.path.join(anomaly_path, entry)
        if os.path.isdir(path):
            if count == 4:
                testList.append(path)
            else :
                trainList.append(path)
            count+=1
    count=0
    for entry in os.listdir(normal_path):
        path = os.path.join(normal_path, entry)
        if os.path.isdir(path):
            if count == 4:
                testList.append(path)
            else :
                trainList.append(path)
            count+=1

    trainListFrames = []
    trainListFramesClass = []
    for folder in trainList:
        className = folder.split('/')[0]
        print(className)
        imgs = os.listdir(folder)
        nbImgs = len(imgs)
        for i in range(nbImgs-NBFRAME):
            train_frames = []
            for j in range(NBFRAME):
                imgPath = folder+'/'+imgs[i+j]
                # loading the image and keeping the target size as (224,224,3)
                img = image.load_img(imgPath, target_size=SIZE+(3,))
                # img.show()
                # converting it to array
                img = image.img_to_array(img)
                # normalizing the pixel value
                img = img/255
                # appending the image to the train_image list
                train_frames.append(img)
            trainListFrames.append(train_frames)
            if className == 'anomaly':
                trainListFramesClass.append([0,1])
            else:
                trainListFramesClass.append([1,0])
    trainListFrames = np.array(trainListFrames)
    trainListFramesClass = np.array(trainListFramesClass)
    print(trainListFrames.shape)
    testListFrames = []
    testListFramesClass = []
    for folder in testList:
        className = folder.split('/')[0]
        print(className)
        imgs = os.listdir(folder)
        nbImgs = len(imgs)
        for i in range(nbImgs-NBFRAME):
            train_frames = []
            for j in range(NBFRAME):
                imgPath = folder+'/'+imgs[i+j]
                # loading the image and keeping the target size as (224,224,3)
                img = image.load_img(imgPath, target_size=(112,112,3))
                # img.show()
                # converting it to array
                img = image.img_to_array(img)
                # normalizing the pixel value
                img = img/255
                # appending the image to the train_image list
                train_frames.append(img)
            testListFrames.append(train_frames)
            if className == 'anomaly':
                testListFramesClass.append([0,1])
            else:
                testListFramesClass.append([1,0])
    
    testListFrames = np.array(testListFrames)
    testListFramesClass = np.array(testListFramesClass)
    print(testListFrames.shape)
    print(testListFramesClass.shape)

    # y_train = np.asarray(trainListFramesClass).astype('float32').reshape((-1,1))
    # y_test = np.asarray(testListFramesClass).astype('float32').reshape((-1,1))
    # print(y_train.shape)
    # print(y_test.shape)

    return trainListFrames, testListFrames, trainListFramesClass, testListFramesClass

def defineTrainTestSetsIMU(NBFRAME = NBFRAME, SIZE = SIZE, CHANNELS=CHANNELS):
    trainList = []
    testList = []

    anomaly_path = 'anomalyIMU/'
    normal_path = 'normalIMU/'
    count=0
    for entry in os.listdir(anomaly_path):
        path = os.path.join(anomaly_path, entry)
        if os.path.isdir(path):
            if count == 4:
                testList.append(path)
            else :
                trainList.append(path)
            count+=1
    count=0
    for entry in os.listdir(normal_path):
        path = os.path.join(normal_path, entry)
        if os.path.isdir(path):
            if count == 4:
                testList.append(path)
            else :
                trainList.append(path)
            count+=1

    trainListSeries = []
    trainListSeriesClass = []
    for folder in trainList:
        className = folder.split('/')[0][:-3]
        print(className)
        data = os.listdir(folder)
        wb = load_workbook(filename = folder+'/' + data[0])
        sheets = wb.worksheets[1:]
        features = []
        for sheet in sheets:
            feature = []
            for val in sheet['B'][:15]:
                feature.append(val.value)
            features.append(feature)
        if className == 'anomaly':
            trainListSeriesClass.append([0,1])
        else:
            trainListSeriesClass.append([1,0])
        trainListSeries.append(features)
    

    trainListSeries = np.asarray(trainListSeries)
    print(trainListSeries.shape)
    print(type(trainListSeries[0]))
    trainListSeriesClass = np.array(trainListSeriesClass)

    testListSeries = []
    testListSeriesClass = []
    for folder in testList:
        className = folder.split('/')[0][:-3]
        print(className)
        data = os.listdir(folder)
        wb = load_workbook(filename = folder+'/' + data[0])
        sheets = wb.worksheets[1:]
        print(sheets)
        if className == 'anomaly':
            testListSeriesClass.append([0,1])
        else:
            testListSeriesClass.append([1,0])
    
    testListSeries = np.array(testListSeries)
    testListSeriesClass = np.array(testListSeriesClass)


    return trainListSeries, testListSeries, trainListSeriesClass, testListSeriesClass

def build_convnet(shape=(112, 112, 3)):
    momentum = .9
    model = keras.Sequential()
    model.add(Conv2D(64, (3,3), input_shape=shape,
        padding='same', activation='relu'))
    model.add(Conv2D(64, (3,3), padding='same', activation='relu'))
    model.add(BatchNormalization(momentum=momentum))
    
    model.add(MaxPool2D())
    
    model.add(Conv2D(128, (3,3), padding='same', activation='relu'))
    model.add(Conv2D(128, (3,3), padding='same', activation='relu'))
    model.add(BatchNormalization(momentum=momentum))
    
    model.add(MaxPool2D())
    
    model.add(Conv2D(256, (3,3), padding='same', activation='relu'))
    model.add(Conv2D(256, (3,3), padding='same', activation='relu'))
    model.add(BatchNormalization(momentum=momentum))
    
    model.add(MaxPool2D())
    
    model.add(Conv2D(512, (3,3), padding='same', activation='relu'))
    model.add(Conv2D(512, (3,3), padding='same', activation='relu'))
    model.add(BatchNormalization(momentum=momentum))
    
    # flatten...
    model.add(GlobalMaxPool2D())
    return model

def action_model(shape=(5, 112, 112, 3), nbout=2):
    # Create our convnet with (112, 112, 3) input shape
    convnet = build_convnet(shape[1:])
    
    # then create our final model
    model = keras.Sequential()
    # add the convnet with (5, 112, 112, 3) shape
    model.add(TimeDistributed(convnet, input_shape=shape))
    # here, you can also use GRU or LSTM
    model.add(GRU(64))
    # and finally, we make a decision network
    model.add(Dense(1024, activation='relu'))
    model.add(Dropout(.5))
    model.add(Dense(512, activation='relu'))
    model.add(Dropout(.5))
    model.add(Dense(128, activation='relu'))
    model.add(Dropout(.5))
    model.add(Dense(64, activation='relu'))
    model.add(Dense(nbout, activation='sigmoid'))
    return model

def autoencoder_modelLSTM(shape=(5, 112, 112, 3), nbout=2):
    

    seq = keras.Sequential()
    seq.add(TimeDistributed(Conv2D(128, (10, 10), strides=2, padding="same"), batch_input_shape=(None,) + shape))
    seq.add(LayerNormalization())
    seq.add(TimeDistributed(Conv2D(64, (6, 6), strides=2, padding="same")))
    seq.add(LayerNormalization())
    # # # # #
    seq.add(ConvLSTM2D(64, (3, 3), padding="same"))
    seq.add(LayerNormalization())
    seq.add(GlobalMaxPool2D())

    seq.add(Dense(256, activation='relu'))
    seq.add(Dropout(.5))
    seq.add(Dense(128, activation='relu'))
    seq.add(Dropout(.5))
    seq.add(Dense(64, activation='relu'))
    seq.add(Dense(nbout, activation='sigmoid'))

    return seq

def train_model():

    X_train, X_test, y_train, y_test = defineTrainTestSets()

    INSHAPE=(NBFRAME,) + SIZE + (CHANNELS,) # (5, 112, 112, 3)
    model = autoencoder_modelLSTM(INSHAPE, 2)
    optimizer = keras.optimizers.Adam(0.001)
    model.compile(
        optimizer,
        'binary_crossentropy',
        metrics=['acc']
    )

    EPOCHS=10
    # create a "chkp" directory before to run that
    # because ModelCheckpoint will write models inside
    callbacks = [
        keras.callbacks.ReduceLROnPlateau(verbose=1),
        keras.callbacks.ModelCheckpoint(
            'chkp_LSTM/weights.{epoch:02d}-{val_loss:.2f}.hdf5',
            verbose=1),
    ]
    model.fit(
        x=X_train,
        y=y_train,
        batch_size=8,
        validation_data=(X_test, y_test),
        verbose=1,
        epochs=EPOCHS,
        callbacks=callbacks
    )

def testSample(folderPath, model):
    kModel = keras.models.load_model(model)
    print(kModel.summary())
    sampleList = []
    imgs = os.listdir(folderPath)
    nbImgs = len(imgs)
    for i in range(nbImgs-NBFRAME):
        sample_frames = []
        for j in range(NBFRAME):
            imgPath = folderPath+'/'+imgs[i+j]
            # loading the image and keeping the target size as (224,224,3)
            img = image.load_img(imgPath, target_size=(112,112,3))
            # img.show()
            # converting it to array
            img = image.img_to_array(img)
            # normalizing the pixel value
            img = img/255
            # appending the image to the train_image list
            sample_frames.append(img)
        sampleList.append(sample_frames)
    sampleList = np.array(sampleList)
    predictions = kModel.predict(sampleList)
    return predictions

defineTrainTestSetsIMU()

# predictionsA = testSample('sampleTestA', 'bestModel.hdf5')
# print(predictionsA)
# print('_________________________________________________')
# predictionsN = testSample('sampleTestN', 'bestModel.hdf5')
# print(predictionsN)