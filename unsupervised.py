
import os
# Force Keras on CPU
# os.environ['CUDA_VISIBLE_DEVICES'] = '-1'

import cv2    
import matplotlib.pyplot as plt
from numpy.core.records import array   
import pandas as pd
from keras.preprocessing import image   
import numpy as np   
from keras.utils import np_utils
from skimage.transform import resize  
from openpyxl import load_workbook

from keras.layers import TimeDistributed, GRU, Dense, Dropout, LayerNormalization, \
    ConvLSTM2D, Lambda, Reshape, Permute, Input, add, Add, concatenate, Conv3D, Conv1D, GaussianNoise

from keras.layers import Conv2D, BatchNormalization, \
    MaxPool2D, MaxPooling2D, UpSampling2D, GlobalMaxPool2D

from keras.models import Model

import keras

import tensorflow as tf
from tensorflow.python.framework.ops import GraphKeys
from tensorflow.python.keras.layers.convolutional import Conv2DTranspose

# Fichier pour faire des tests d'apprentissages supervisés, seulement sur les images pour l'instant


print(tf.test.is_built_with_cuda())
print(tf.config.list_physical_devices())

# nombre de frames par séquence
NBFRAME = 6
# taille des images après redimensionnement pour fit
SIZE = (112, 112)
# TODO : convertir image en greyscale pour entrainement (done)
# TODO : save image in (112, 112) to load faster
# TODO : data augmentation

# création des séquences d'images pour entrainement et validation

def defineTrainSets(NBFRAME = NBFRAME, SIZE = SIZE, grayscale = False): 
    trainList = []
    testList = []

    normal_path = 'normal/'
    count=0
    for entry in os.listdir(normal_path):
        path = os.path.join(normal_path, entry)
        if os.path.isdir(path):
            # on garde le cinquième dossier pour les tests
            if count == 4:
                testList.append(path)
            else :
                trainList.append(path)
            count+=1

    trainListFrames = []
    for folder in trainList:
        className = folder.split('/')[0]
        print(className)
        imgs = os.listdir(folder)
        nbImgs = len(imgs)
        for i in range(nbImgs-NBFRAME):
            train_frames = []
            for j in range(NBFRAME):
                imgPath = folder+'/'+imgs[i+j]
                if grayscale:
                    img = image.load_img(imgPath, color_mode="grayscale", target_size=SIZE+(1,))
                else:
                    img = image.load_img(imgPath, target_size=SIZE+(3,))
                img = image.img_to_array(img)
                img = img/255
                train_frames.append(img)
            trainListFrames.append(train_frames)
            
    trainListFrames = np.array(trainListFrames)

    return trainListFrames

def defineTrainSetsNextFramePred(NBFRAME = NBFRAME, SIZE = SIZE, grayscale = False): 
    trainList = []
    testList = []

    normal_path = 'normal/'
    count=0
    for entry in os.listdir(normal_path):
        path = os.path.join(normal_path, entry)
        if os.path.isdir(path):
            # on garde le cinquième dossier pour les tests
            if count == 4:
                testList.append(path)
            else :
                trainList.append(path)
            count+=1

    trainListFrames = []
    trainListPreds = []
    for folder in trainList:
        className = folder.split('/')[0]
        print(className)
        imgs = os.listdir(folder)
        nbImgs = len(imgs)
        for i in range(nbImgs-NBFRAME):
            train_frames = []
            for j in range(NBFRAME):
                imgPath = folder+'/'+imgs[i+j]
                if grayscale:
                    img = image.load_img(imgPath, color_mode="grayscale", target_size=SIZE+(1,))
                else:
                    img = image.load_img(imgPath, target_size=SIZE+(3,))
                img = image.img_to_array(img)
                img = img/255
                if j == NBFRAME-1:
                    trainListPreds.append([img])
                else:
                    train_frames.append(img)
            trainListFrames.append(train_frames)
    trainListFrames = np.array(trainListFrames)
    trainListPreds = np.array(trainListPreds)
    print(trainListFrames.shape)
    print(trainListPreds.shape)

    testListFrames = []
    testListPreds = []
    for folder in testList:
        className = folder.split('/')[0]
        print(className)
        imgs = os.listdir(folder)
        nbImgs = len(imgs)
        for i in range(nbImgs-NBFRAME):
            test_frames = []
            for j in range(NBFRAME):
                imgPath = folder+'/'+imgs[i+j]
                if grayscale:
                    img = image.load_img(imgPath, color_mode="grayscale", target_size=SIZE+(1,))
                else:
                    img = image.load_img(imgPath, target_size=SIZE+(3,))
                img = image.img_to_array(img)
                img = img/255
                if j == NBFRAME-1:
                    testListPreds.append([img])
                else :
                    test_frames.append(img)
            testListFrames.append(test_frames)

    testListFrames = np.array(testListFrames)
    testListPreds = np.array(testListPreds)
    print(testListFrames.shape)
    print(testListPreds.shape)

    return trainListFrames, testListFrames, trainListPreds, testListPreds

# création des séquences de données IMU pour entrainement et validation (pas fini!)
def defineTrainTestSetsIMU(NBFRAME = NBFRAME, SIZE = SIZE):
    trainList = []
    testList = []

    anomaly_path = 'anomalyIMU/'
    normal_path = 'normalIMU/'
    count=0
    for entry in os.listdir(anomaly_path):
        path = os.path.join(anomaly_path, entry)
        if os.path.isdir(path):
            if count == 4:
                testList.append(path)
            else :
                trainList.append(path)
            count+=1
    count=0
    for entry in os.listdir(normal_path):
        path = os.path.join(normal_path, entry)
        if os.path.isdir(path):
            if count == 4:
                testList.append(path)
            else :
                trainList.append(path)
            count+=1

    trainListSeries = []
    trainListSeriesClass = []
    for folder in trainList:
        className = folder.split('/')[0][:-3]
        print(className)
        data = os.listdir(folder)
        wb = load_workbook(filename = folder+'/' + data[0])
        sheets = wb.worksheets[1:]
        features = []
        for sheet in sheets:
            feature = []
            for val in sheet['B'][:15]:
                feature.append(val.value)
            features.append(feature)
        if className == 'anomaly':
            trainListSeriesClass.append([0,1])
        else:
            trainListSeriesClass.append([1,0])
        trainListSeries.append(features)
    

    trainListSeries = np.asarray(trainListSeries)
    trainListSeriesClass = np.array(trainListSeriesClass)

    testListSeries = []
    testListSeriesClass = []
    for folder in testList:
        className = folder.split('/')[0][:-3]
        print(className)
        data = os.listdir(folder)
        wb = load_workbook(filename = folder+'/' + data[0])
        sheets = wb.worksheets[1:]
        features = []
        for sheet in sheets:
            feature = []
            for val in sheet['B'][:15]:
                feature.append(val.value)
            features.append(feature)
        if className == 'anomaly':
            testListSeriesClass.append([0,1])
        else:
            testListSeriesClass.append([1,0])
    
    testListSeries = np.array(testListSeries)
    testListSeriesClass = np.array(testListSeriesClass)


    return trainListSeries, testListSeries, trainListSeriesClass, testListSeriesClass

# build_convnet_encoder et build_convnet_decoder ne sont pas utilisées pour l'instant,
# utilisation d'un autre autoencoder à la place
def build_convnet_encoder(shape=(112, 112, 3)):
    momentum = .9
    model = keras.Sequential()
    model.add(Conv2D(128, (3,3), input_shape=shape,
        padding='same', activation='relu'))
    model.add(BatchNormalization(momentum=momentum))
    
    model.add(MaxPool2D())
    model.add(Conv2D(64, (3,3), padding='same', activation='relu'))
    model.add(BatchNormalization(momentum=momentum))
    
    # flatten...
    model.add(GlobalMaxPool2D())
    return model

def build_convnet_decoder():
    momentum = .9
    model = keras.Sequential()
    model.add(Conv2DTranspose(64, (3,3), padding='same', activation='relu'))
    model.add(BatchNormalization(momentum=momentum))
    
    model.add(MaxPool2D())
    model.add(Conv2DTranspose(128, (3,3), padding='same', activation='relu'))
    model.add(BatchNormalization(momentum=momentum))
    
    # flatten...
    # model.add(GlobalMaxPool2D())
    return model

# autoencoder TimeDistributed Conv2D et ConvLSTM2D, 
def autoencoder_model(shape=(5, 112, 112, 1)):
    

    seq = keras.Sequential()
    seq.add(TimeDistributed(Conv2D(128, (10, 10), strides=2, padding="same"), batch_input_shape=(None,) + shape))
    seq.add(LayerNormalization())
    seq.add(TimeDistributed(Conv2D(64, (6, 6), strides=2, padding="same")))
    seq.add(LayerNormalization())
    # # # # #
    seq.add(ConvLSTM2D(64, (3, 3), padding="same", return_sequences=True))
    seq.add(LayerNormalization())
    seq.add(ConvLSTM2D(32, (3, 3), padding="same", return_sequences=True))
    seq.add(LayerNormalization())
    seq.add(ConvLSTM2D(64, (3, 3), padding="same", return_sequences=True))
    seq.add(LayerNormalization())
    # # # # #
    seq.add(TimeDistributed(Conv2DTranspose(128, (6, 6), strides=2, padding="same")))
    seq.add(LayerNormalization())
    seq.add(TimeDistributed(Conv2DTranspose(1, (11, 11), strides=2, padding="same")))
    # seq.add(LayerNormalization())
    # seq.add(TimeDistributed(Conv2D(3, (11, 11), activation="sigmoid", padding="same")))



    return seq

# test d'un autre autoencoder, non fonctionnel pour l'instant
def autoencoder_modelV2(shape=(5, 112, 112, 3)):
    # Create our convnet with (112, 112, 3) input shape
    convnet_encoder = build_convnet_encoder(shape[1:])
    
    # then create our final model
    model = keras.Sequential()

    model.add(TimeDistributed(convnet_encoder, input_shape = shape))

    # add the convnet with (5, 112, 112, 3) shape
    # model.add(TimeDistributed(convnet_encoder, input_shape=shape))
    # here, you can also use GRU or LSTM
    model.add(GRU(64, return_sequences=True))
    model.add(GRU(32, return_sequences=True))
    model.add(GRU(64, return_sequences=True))
    convnet_decoder = build_convnet_decoder()
    model.add(TimeDistributed(convnet_decoder))
    model.add(TimeDistributed(Conv2D(3, (11, 11), activation="sigmoid", padding="same")))
    return model

# modèle de prédiction de frame semi-supervisé, simple mais bon résultat
def next_frame_pred_model_lightweight(shape=(5, 112, 112, 1)):
    
    def slice(x):
        return x[:,:,:,:, -1]


    inp = Input(shape)
    permuted = Permute((2,3,4,1))(inp)
    noise = GaussianNoise(0.1)(permuted)
    last_layer = Lambda(slice, input_shape=(112,112,1,5), output_shape=(112,112,1))(noise)
    permuted_2 = Permute((4,1,2,3))(noise)

    conv_lstm_output_1 = ConvLSTM2D(6, (3,3), padding='same')(permuted_2)
    conv_output = Conv2D(1, (3,3), padding="same")(conv_lstm_output_1)
    combined = add([last_layer, conv_output])

    model=Model(inputs=[inp], outputs=[combined])

    return model

# modèle de prédiction de frame semi-supervisé, plus compliqué, difficile de savoir si les résultats obtenus sont meilleurs que les précédents
def next_frame_pred_model(shape, grayscale=False):
    channels = 3
    if grayscale: 
        channels = 1
    
    def slice(x):
        return x[:,:,:,:, -1]
    
    def cut(x):
        return x[:,:,:,:, :-1]
    c = 4

    inp = Input(shape)
    # cut_last_layer = Lambda(cut, input_shape=(112,112,1,6), output_shape=(112,112,1,5))(inp)
    permuted = Permute((2,3,4,1))(inp)
    noise = GaussianNoise(0.1)(permuted)
    last_layer = Lambda(slice, input_shape=(112,112,channels,5), output_shape=(112,112,channels))(noise)
    x = Permute((4,1,2,3))(noise)

    x =(ConvLSTM2D(filters=c, kernel_size=(3,3),padding='same',name='conv_lstm1', return_sequences=True))(x)

    c1=(BatchNormalization())(x)
    x = Dropout(0.2)(x)
    x =(TimeDistributed(MaxPooling2D(pool_size=(2,2))))(c1)

    x =(ConvLSTM2D(filters=2*c,kernel_size=(3,3),padding='same',name='conv_lstm3',return_sequences=True))(x)
    c2=(BatchNormalization())(x)
    x = Dropout(0.2)(x)

    x =(TimeDistributed(MaxPooling2D(pool_size=(2,2))))(c2)
    x =(ConvLSTM2D(filters=4*c,kernel_size=(3,3),padding='same',name='conv_lstm4',return_sequences=True))(x)

    x =(TimeDistributed(UpSampling2D(size=(2, 2))))(x)
    x =(ConvLSTM2D(filters=4*c,kernel_size=(3,3),padding='same',name='conv_lstm5',return_sequences=True))(x)
    x =(BatchNormalization())(x)

    x =(ConvLSTM2D(filters=2*c,kernel_size=(3,3),padding='same',name='conv_lstm6',return_sequences=True))(x)
    x =(BatchNormalization())(x)
    x = Add()([c2, x])
    x = Dropout(0.2)(x)

    x =(TimeDistributed(UpSampling2D(size=(2, 2))))(x)
    x =(ConvLSTM2D(filters=c,kernel_size=(3,3),padding='same',name='conv_lstm7',return_sequences=False))(x)
    x =(BatchNormalization())(x)
    combined = concatenate([last_layer, x])
    combined = Conv2D(channels, (1,1))(combined)
    model=Model(inputs=[inp], outputs=[combined])

    return model

# modèle de prédiction de frame semi-supervisé, encore plus compliqué, ne fonctionne pas correctement pour l'instant
def next_frame_pred_model_experimental(shape, grayscale=False):
    channels = 3
    if grayscale: 
        channels = 1
    
    def slice(x):
        return x[:,:,:,:, -1]
    
    def cut(x):
        return x[:,:,:,:, :-1]
    c = 4

    inp = Input(shape)
    # cut_last_layer = Lambda(cut, input_shape=(112,112,1,6), output_shape=(112,112,1,5))(inp)
    permuted = Permute((2,3,4,1))(inp)
    # last_layer = Lambda(slice, input_shape=(112,112,channels,5), output_shape=(112,112,channels))(noise)
    x = Permute((4,1,2,3))(permuted)

    x =(ConvLSTM2D(filters=5, kernel_size=(3,3),padding='same',name='conv_lstm1', return_sequences=True))(x)
    x =(ConvLSTM2D(filters=5, kernel_size=(3,3),padding='same',name='conv_lstm2', return_sequences=True))(x)
    lstm3 =(ConvLSTM2D(filters=5, kernel_size=(3,3),padding='same',name='conv_lstm3', return_sequences=True))(x)
    x =(TimeDistributed(MaxPooling2D(pool_size=(2,2))))(lstm3)
    x =(ConvLSTM2D(filters=10, kernel_size=(3,3),padding='same',name='conv_lstm4', return_sequences=True))(x)
    x =(ConvLSTM2D(filters=10, kernel_size=(3,3),padding='same',name='conv_lstm5', return_sequences=True))(x)
    lstm6 =(ConvLSTM2D(filters=10, kernel_size=(3,3),padding='same',name='conv_lstm6', return_sequences=True))(x)
    x =(TimeDistributed(MaxPooling2D(pool_size=(2,2))))(lstm6)
    x =(ConvLSTM2D(filters=10, kernel_size=(3,3),padding='same',name='conv_lstm7', return_sequences=True))(x)
    x =(ConvLSTM2D(filters=10, kernel_size=(3,3),padding='same',name='conv_lstm8', return_sequences=True))(x)
    x =(ConvLSTM2D(filters=10, kernel_size=(3,3),padding='same',name='conv_lstm9', return_sequences=True))(x)
    x =(TimeDistributed(UpSampling2D(size=(2, 2))))(x)
    combined = concatenate([lstm6, x])
    x =(TimeDistributed(Conv2D(filters = 5, kernel_size=(3,3), padding="same")))(combined)
    x =(TimeDistributed(UpSampling2D(size=(2, 2))))(x)
    combined = concatenate([lstm3, x])
    x =(TimeDistributed(Conv2D(filters = 3, kernel_size=(3,3), padding="same")))(combined)
    x =(TimeDistributed(UpSampling2D(size=(2, 2))))(x)
    x =(TimeDistributed(Conv2D(filters = 3, kernel_size=(3,3), padding="same")))(x)
    x =(TimeDistributed(MaxPooling2D(pool_size=(2,2))))(x)
    x = (Conv3D(filters=3, kernel_size=(3,3,3), padding="same"))(x)
    x = Permute((2,3,4,1))(x)

    x = Conv1D(channels, (3))(x)
    x = Conv1D(channels, (1))(x)
    model=Model(inputs=[inp], outputs=[x])
    return model


# fonction à appeler pour entraîner un model d'autoencoder, entraîne uniquement sur les données normales (semi-supervised)
# (à modifier pour rendre plus modulable pour l'instant il faut modifier la fnction si on veut entraîner un autre model)
def train_model():

    X_train = defineTrainSets()

    INSHAPE=(NBFRAME,) + SIZE + (1,) # (5, 112, 112, 1)
    model = autoencoder_model(INSHAPE)

    model.compile(loss='mse', optimizer=keras.optimizers.Adam(lr=1e-4, decay=1e-5, epsilon=1e-6))

    EPOCHS=50
    # create a "chkp" directory before to run that
    # because ModelCheckpoint will write models inside
    callbacks = [
        keras.callbacks.ReduceLROnPlateau(verbose=1),
        keras.callbacks.ModelCheckpoint(
            'unsupervised_model_normal_only/weights.{epoch:02d}.hdf5',
            verbose=1),
    ]
    model.fit(
        x=X_train,
        y=X_train,
        batch_size=8,
        verbose=1,
        epochs=EPOCHS,
        callbacks=callbacks
    )


# métrique utiliser pour l'évaluation des modèles de prédictions de frame suivant
def nextFramePredMetric(y_true, y_pred):
    g = y_true[:, :, :, 0] - y_pred[:, :, :, 0]
    
    return keras.backend.mean(keras.backend.sqrt(g*g))

# fonction à appeler pour entraîner un model de prédiction de frame, entraîne uniquement sur les données normales (semi-supervised)
# (à modifier pour rendre plus modulable pour l'instant il faut modifier la fnction si on veut entraîner un autre model)
def train_model_next_frame_pred(grayscale = False):

    CHANNELS = 3
    if grayscale:
        CHANNELS = 1

    X_train, X_test, y_train, y_test = defineTrainSetsNextFramePred(grayscale=grayscale)

    INSHAPE=(NBFRAME-1,) + SIZE + (CHANNELS,) # (5, 112, 112, 1)
    model = autoencoder_next_frame_pred_model(INSHAPE, grayscale = grayscale)

    model.compile(optimizer='adam', loss='mse', metrics=[nextFramePredMetric])


    EPOCHS=100
    # create a "chkp" directory before to run that
    # because ModelCheckpoint will write models inside
    callbacks = [
        keras.callbacks.ReduceLROnPlateau(verbose=1),
        keras.callbacks.ModelCheckpoint(
            'next_frame_pred_model_normal_only_lightweight/weights.{epoch:02d}.hdf5',
            verbose=1),
    ]
    model.fit(
        x=X_train,
        y=y_train,
        batch_size=8,
        validation_data=(X_test, y_test),
        verbose=1,
        epochs=EPOCHS,
        callbacks=callbacks
    )

# fonction à appeler pour faire des tests avec un model d'autoencoder enregistré, sur une série de séquence
# enregistre les images prédites, et renvoie un graphe de la régularité des prédictions au fil des séquences,
# détecte au moins une anomalie dans une suite de séquences d'anomalies,
# mais classifie une partie des séquences des anomalies comme normales
# est-ce le comportement voulu? (on trouve bien une anomalie dans chaque "vidéo" d'anomalie)
def testSample(folderPath, model, grayscale = False):
    kModel = keras.models.load_model(model)
    sampleList = []
    imgs = os.listdir(folderPath)
    nbImgs = len(imgs)
    for i in range(nbImgs-NBFRAME):
        sample_frames = []
        for j in range(NBFRAME):
            imgPath = folderPath+'/'+imgs[i+j]
            if grayscale:
                img = image.load_img(imgPath, color_mode="grayscale", target_size=(112,112,1))
            else:
                img = image.load_img(imgPath, target_size=(112,112,3))
            # img.show()
            # converting it to array
            img = image.img_to_array(img)
            # normalizing the pixel value
            img = img/255
            if not os.path.isdir("./original_imgs/" + folderPath):
                os.makedirs("./original_imgs/" + folderPath)
            image.save_img("./original_imgs/" + folderPath+'/' +str(i)+'_'+str(j)+'.jpeg', img)
            # appending the image to the train_image list
            sample_frames.append(img)
        sampleList.append(sample_frames)
    sampleList = np.array(sampleList)
    predictions = kModel.predict(sampleList, batch_size = 4)
    for i,seq in enumerate(predictions):
        for j,img in enumerate(seq):
            if not os.path.isdir("./predicted_imgs/" + folderPath):
                os.makedirs("./predicted_imgs/" + folderPath)
            image.save_img("./predicted_imgs/" + folderPath +'/'+ str(i)+'_'+str(j)+'.jpeg', img)
    n=len(predictions)
    sequences_reconstruction_cost = np.array([np.linalg.norm(np.subtract(sampleList[i],predictions[i])) for i in range(0,n)])
    sa = (sequences_reconstruction_cost - np.min(sequences_reconstruction_cost)) / np.max(sequences_reconstruction_cost)
    sr = 1.0 - sa

    # plot the regularity scores
    plt.plot(sr)
    plt.ylabel('regularity score Sr(t)')
    plt.xlabel('frame t')
    plt.show()
    return predictions


# même idée que la fonction précédente, mais avec des modèles de prédiction de frame
def testSampleNextFramePred(folderPath, kModel, grayscale = False, ax = plt, label = ""):
    sampleList = []
    testList = []
    imgs = os.listdir(folderPath)
    nbImgs = len(imgs)
    for i in range(nbImgs-NBFRAME):
        sample_frames = []
        for j in range(NBFRAME):
            imgPath = folderPath+'/'+imgs[i+j]
            if grayscale:
                img = image.load_img(imgPath, color_mode="grayscale", target_size=(112,112,1))
            else:
                img = image.load_img(imgPath, target_size=(112,112,3))
            # img.show()
            # converting it to array
            img = image.img_to_array(img)
            # normalizing the pixel value
            img = img/255
            if not os.path.isdir("./original_imgs/" + folderPath):
                os.makedirs("./original_imgs/" + folderPath)
            if j == NBFRAME -1 :
                image.save_img("./original_imgs/" + folderPath+'/' +str(i)+'_'+str(j)+ '_to_pred' + '.jpeg', img)
                testList.append(img)

            else:
                image.save_img("./original_imgs/" + folderPath+'/' +str(i)+'_'+str(j)+ '.jpeg', img)
                # appending the image to the train_image list
                sample_frames.append(img)
        sampleList.append(sample_frames)
    testList = np.array(testList)
    sampleList = np.array(sampleList)
    predictions = kModel.predict(sampleList, batch_size = 4)
    predictions = predictions[:, :, :, :, 0]
    print(predictions.shape)
    print(sampleList.shape)
    for i,img in enumerate(predictions):
        if not os.path.isdir("./predicted_imgs/" + folderPath):
            os.makedirs("./predicted_imgs/" + folderPath)
        image.save_img("./predicted_imgs/" + folderPath +'/'+ str(i)+'.jpeg', img)
    n=len(predictions)
    sequences_reconstruction_cost = np.array([np.linalg.norm(np.subtract(testList[i],predictions[i])) for i in range(0,n)])
    sa = (sequences_reconstruction_cost - np.min(sequences_reconstruction_cost)) / np.max(sequences_reconstruction_cost)
    sr = 1.0 - sa

    # plot the regularity scores
    ax.plot(sr, label = label)
    ax.set_ylabel('regularity score Sr(t)')
    ax.set_xlabel('frame t')
    # ax.show()
    return predictions

# expérimentation avec l'optical flow pour utliser en données d'entraînement, pas testé encore,
# mais données obtenues pas très propres
def optical_flow(folderPath):
    imgs = os.listdir(folderPath)
    nbImgs = len(imgs)
    for i in range(nbImgs-1):
        sample_frames = []
        imgPath = folderPath+'/'+imgs[i]
        imgprev = cv2.imread(imgPath, 1)
        hsv = np.zeros_like(imgprev)
        hsv[...,1] = 255
        imgprev = cv2.cvtColor(imgprev,cv2.COLOR_BGR2GRAY)
        imgPath = folderPath+'/'+imgs[i+1]
        imgnext = cv2.imread(imgPath, 1)
        imgnext = cv2.cvtColor(imgnext,cv2.COLOR_BGR2GRAY)
        

        flow = cv2.calcOpticalFlowFarneback(imgprev,imgnext, None, 0.5, 3, 15, 3, 5, 1.2, 0)
        mag, ang = cv2.cartToPolar(flow[...,0], flow[...,1])
        hsv[...,0] = ang*180/np.pi/2
        hsv[...,2] = cv2.normalize(mag,None,0,255,cv2.NORM_MINMAX)
        rgb = cv2.cvtColor(hsv,cv2.COLOR_HSV2BGR)
        if not os.path.isdir("./optical_flow/" + folderPath):
                os.makedirs("./optical_flow/" + folderPath)
        cv2.imwrite("./optical_flow/" + folderPath+'/' +str(i)+'.jpeg', rgb)

# fonction pour tester la "régularité" d'un modèle de prédiction de frame sur tout le dataset
def benchmark_all_dataset(model, axs, label):
    testSampleNextFramePred('normal/2020-01-17-11-32-12', model, grayscale=True, ax=axs[0,0], label = label)
    testSampleNextFramePred('normal/2020-01-17-11-32-49', model, grayscale=True, ax=axs[0,1], label = label)
    testSampleNextFramePred('normal/2020-01-17-11-33-26', model, grayscale=True, ax=axs[0,2], label = label)
    testSampleNextFramePred('normal/2020-01-17-11-34-08', model, grayscale=True, ax=axs[0,3], label = label)
    testSampleNextFramePred('normal/2020-01-17-11-34-43', model, grayscale=True, ax=axs[0,4], label = label)

    testSampleNextFramePred('anomaly/2020-01-17-11-35-27', model, grayscale=True, ax=axs[1,0], label = label)
    testSampleNextFramePred('anomaly/2020-01-17-11-36-03', model, grayscale=True, ax=axs[1,1], label = label)
    testSampleNextFramePred('anomaly/2020-01-17-11-36-43', model, grayscale=True, ax=axs[1,2], label = label)
    testSampleNextFramePred('anomaly/2020-01-17-11-37-25', model, grayscale=True, ax=axs[1,3], label = label)
    testSampleNextFramePred('anomaly/2020-01-17-11-38-07', model, grayscale=True, ax=axs[1,4], label = label)


